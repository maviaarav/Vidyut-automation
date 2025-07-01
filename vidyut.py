from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openai import OpenAI
import base64
import time
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import os
# ...existing code...
# ...existing code...
registered_mobile = str(input("Enter Mobile no: "))
password = str(input("Enter Password: "))
is_phoneno_same = input("Is Agent's phone number same? (Yes/No): ").capitalize()
if is_phoneno_same == 'Yes':
    Phoneno = registered_mobile
else:
    Phoneno = input("Enter alternate phone number: ")
House_no = input("Enter house number: ")
Building = input("Enter building/Colony: ")
Landmark = input("Enter landmark: ")
Locality = input("Enter locality: ")
Pincode = input("Enter pincode: ")
AgentName = input("Enter agent name: ")
MailId = input("Enter mail id: ")
is_new_lift_or_escalator = input("Is new lift or escalator? (Yes/No): ")
print('''
        Premise Type:
      Press the following number keys written in front of the Options:
      Commercial = 1
      Centeral Government = 2
      State Goverment = 3
      PSU = 4
      Housing Society = 5
      Industrial = 6
      Others = 7
     ''')
Premise_type = int(input("Enter the Number based on the above infromation"))
IsPublicORPrivatePremise = input("Is premise Public or Private?: ").capitalize()
OC = input("Is OC available? (Yes/No): ").capitalize()
if OC == "Yes":
    Authority = input("Enter authority name: ")
    ApprovalNumber = input("Enter approval number: ")
    DateOfApproval = input("Enter date of approval (dd-mm-yyyy): ")
confirmation = input("Is Your Appliance a Lift? (Yes/No): ").capitalize()
if confirmation == 'Yes':
    Applience = "Lift"
else:
    Applience = "Escalator"
Make = input("Enter make: ")
weight = input("Enter weight: ")
Person_cap = float(weight) / 68
agency_house_no = input("Enter agency house number: ")
agency_building = input("Enter agency building/colony: ")
agency_landmark = input("Enter agency landmark: ")
agency_locality = input("Enter agency locality: ")
agency_pincode = input("Enter agency pincode: ")
local_authorized_representative = input("Enter local authorized representative: ")
commence_date = input("Enter commencement date (yyyy-mm-dd): ")
end_date = input("Enter end date (yyyy-mm-dd): ")
OC_file = input(r"Enter OC file path: ")
Drawing = input(r"Enter drawing file path: ")
Affidavit_Manufacturer = input(r"Enter affidavit manufacturer file path: ")
Technical_detail = input(r"Enter technical detail file path: ")
Safety = input(r"Enter safety file path: ")
Signed_by_three_Affidavit = input(r"Enter signed by three affidavit file path: ")
Owners_Signature = input(r"Enter owner's signature file path: ")
Manufacturer_Signature = input(r"Enter manufacturer's signature file path: ")
how_many_form = int(input("How many forms?: "))
print(f" You mean {how_many_form} registration to be filled...")
client = OpenAI(api_key="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")

def extract_captcha():
    with open("Captcha.png", "rb") as image_file:
        base64_image = base64.b64encode(image_file.read()).decode('utf-8')

    response = client.chat.completions.create(
        model="gpt-4.1",
        temperature=0.2,
        messages=[
            {"role": "system", "content": "You are a helpful assistant that extracts captcha codes from images. Just return the captcha code as text. not any other information."},
            {
                "role": "user",
                "content": [
                    {"type": "text", "text": "Return only the captcha code shown in the image, no explanation."},
                    {"type": "image_url", "image_url": {"url": f"data:image/png;base64,{base64_image}"}}
                ]
            }
        ],
        max_tokens=20
    )
    return response.choices[0].message.content.strip()

options = Options()
options.add_argument("--headless")
options.add_experimental_option("prefs", {
    "credentials_enable_service": False,
    "profile.password_manager_enabled": False
})

devicer = webdriver.Chrome(options=options)
devicer.get("https://updeslift.org/Account/login")
devicer.set_window_size(1920, 3000)

def login():
    for attempt in range(5):
        print(f"\nAttempt {attempt + 1} to login with captcha.")
        captcha = devicer.find_element(By.XPATH, '//*[@id="loginform"]/div[5]')
        captcha.screenshot('Captcha.png')
        captcha_code = extract_captcha()
        print("Captcha Code:", captcha_code)

        Type = devicer.find_element(By.CSS_SELECTOR, "select#Type")
        Select(Type).select_by_index(1)
        devicer.find_element(By.ID, "Mob").clear()
        devicer.find_element(By.ID, "Mob").send_keys(registered_mobile)
        devicer.find_element(By.ID, "Password").clear()
        devicer.find_element(By.ID, "Password").send_keys(password)
        devicer.find_element(By.ID, "Captcha").clear()
        devicer.find_element(By.ID, "Captcha").send_keys(captcha_code)

        devicer.find_element(By.CSS_SELECTOR, "button.btn.btn-primary.btn-lg").click()
        print("Submitted login form.")

        time.sleep(2)

        try:
            WebDriverWait(devicer, 2).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "li.nav-parent"))
            )
            print("Login successful!")
            break
        except:
            print("Login failed, probably due to wrong captcha.")
            if attempt == 4:
                print("Failed to login after 5 attempts.")
                devicer.quit()
                return

    # Form filling begins here
    i = 0
    forms = 0
    models = ["KONE111", "OTIS220", "SCHINDLER330", "HYUNDAI440", "JOHNSON550"]
    while i <=how_many_form:
        devicer.find_element(By.CSS_SELECTOR, "li.nav-parent").click()
        devicer.find_element(By.CSS_SELECTOR, "a[href='/Admin/ListAnnexure1']").click()
        devicer.find_element(By.CSS_SELECTOR, "a.btn.btn-success").click()

        devicer.find_element(By.ID, "Annexure1s_OwnerLocalHouseNo").send_keys(House_no)
        devicer.find_element(By.ID, "Annexure1s_OwnerLocalBuildingNo").send_keys(Building)
        devicer.find_element(By.ID, "Annexure1s_OwnerLocalLandmark").send_keys(Landmark)
        devicer.find_element(By.ID, "Annexure1s_OwnerLocalLocality").send_keys(Locality)
        devicer.find_element(By.ID, "Annexure1s_Owner_Local_Pincode").send_keys(Pincode, Keys.TAB)
        devicer.find_element(By.ID, "Annexure1s_OwnerPermanentHouseNo").send_keys(House_no)
        devicer.find_element(By.ID, "Annexure1s_OwnerPermanentBuilding").send_keys(Building)
        devicer.find_element(By.ID, "Annexure1s_OwnerPermanentLandmark").send_keys(Landmark)
        devicer.find_element(By.ID, "Annexure1s_OwnerPermanentLocality").send_keys(Locality)
        devicer.find_element(By.ID, "Annexure1s_OwnerPermanentPincode").send_keys(Pincode,Keys.TAB)
        time.sleep(1)
        devicer.find_element(By.ID, "nxt1").click()
        time.sleep(2)
        devicer.save_screenshot("Step1.png")
        print("Form 1 Filled Successfully.")

        devicer.find_element(By.ID, "Annexure1s_AgentName").send_keys(AgentName)
        devicer.find_element(By.ID, "Annexure1s_AgentLocalHouseNo").send_keys(House_no)
        devicer.find_element(By.ID, "Annexure1s_AgentLocalBuildingNo").send_keys(Building)
        devicer.find_element(By.ID, "Annexure1s_AgentLocalLandmark").send_keys(Landmark)
        devicer.find_element(By.ID, "Annexure1s_AgentLocalLocality").send_keys(Locality)
        devicer.find_element(By.ID, "Annexure1s_Agent_Local_Pincode").send_keys(Pincode, Keys.TAB)
        devicer.find_element(By.ID, "Annexure1s_AgentPermanentHouseNo").send_keys(House_no)
        devicer.find_element(By.ID, "Annexure1s_AgentPermanentBuilding").send_keys(Building)
        devicer.find_element(By.ID, "Annexure1s_AgentPermanentLandmark").send_keys(Landmark)
        devicer.find_element(By.ID, "Annexure1s_AgentPermanentLocality").send_keys(Locality)
        devicer.find_element(By.ID, "Annexure1s_AgentPermanentPincode").send_keys(Pincode, Keys.TAB)
        devicer.find_element(By.ID, "Annexure1s_AgentMailId").send_keys(MailId)
        devicer.find_element(By.ID, "Annexure1s_AgentMob").send_keys(Phoneno)
        time.sleep(2)
        devicer.find_element(By.ID, "nxt2").click()

        devicer.save_screenshot("Step2.png")
        print("Form 2 Filled Successfully.")

        value = "Yes" if is_new_lift_or_escalator == "Yes" else "No"
        radio = devicer.find_element(By.XPATH, f"//*[@id='IsnewLift' and @value='{value}']")
        devicer.execute_script("arguments[0].scrollIntoView(true);", radio)
        time.sleep(1)
        devicer.execute_script("arguments[0].click();", radio)

        devicer.find_element(By.ID, "AnnexxIIs_PremiseHouseNo").send_keys(House_no)
        devicer.find_element(By.ID, "AnnexxIIs_PremiseBuildingNo").send_keys(Building)
        devicer.find_element(By.ID, "AnnexxIIs_PremiseLandmark").send_keys(Landmark)
        devicer.find_element(By.ID, "AnnexxIIs_PremiseLocality").send_keys(Locality)
        devicer.find_element(By.ID, "AnnexxIIs_Premise_Pincode").send_keys(Pincode, Keys.TAB)
        time.sleep(2)
        Select(devicer.find_element(By.CSS_SELECTOR, "select#AnnexxIIs_IsPublicORPrivatePremise")).select_by_visible_text(IsPublicORPrivatePremise)
        Select(devicer.find_element(By.CSS_SELECTOR, "select#AnnexxIIs_PremiseType")).select_by_index(Premise_type)
        devicer.find_element(By.XPATH, "//*[@id='IsLiftModifyOrAltered' and @value = 'No']").click()
        devicer.find_element(By.XPATH, "//*[@id='IsLiftShifted' and @value = 'No']").click()

        if OC == "yes":
            devicer.find_element(By.XPATH, "//*[@id='IsBuildingMapApproved' and @value = 'Yes']").click()
            devicer.find_element(By.ID, "AnnexxIIs_AuthorityApprovingName").send_keys(Authority)
            devicer.find_element(By.ID, "AnnexxIIs_ApprovalNo").send_keys(ApprovalNumber)
            devicer.find_element(By.ID, "AnnexxIIs_ApprovalDate").send_keys(DateOfApproval)
        else:
            devicer.find_element(By.XPATH, "//*[@id='IsBuildingMapApproved' and @value = 'No']").click()

        time.sleep(3)
        print("Form 3 Successfully filled.")
        devicer.save_screenshot("Step3.png")
        devicer.find_element(By.ID, "nxt3").click()
        time.sleep(2)


        if Applience == "Lift":
            radio = devicer.find_element(By.XPATH, "//*[@id='LiftType' and @value='Lift']")
            devicer.execute_script("arguments[0].click();", radio)
            devicer.find_element(By.XPATH, "//*[@id='LiftType' and @value = 'Lift']").click()
            devicer.find_element(By.CSS_SELECTOR, "input#AnnexIV_Make").send_keys(Make)
            model_name = models[i % len(models)]  # cycle through models if i exceeds list
            devicer.find_element(By.CSS_SELECTOR, "input#AnnexIV_Model").send_keys(model_name)

        else:
            radio = devicer.find_element(By.XPATH, "//*[@id='LiftType' and @value='Escalator']")
            devicer.execute_script("arguments[0].click();", radio)
            devicer.find_element(By.CSS_SELECTOR, "input#AnnexIV_Make").send_keys(Make)
            model_name = models[i % len(models)]  
            devicer.find_element(By.CSS_SELECTOR, "input#AnnexIV_Model").send_keys(model_name)

        if is_new_lift_or_escalator == "Yes":
            Select(devicer.find_element(By.CSS_SELECTOR, "select#AnnexIV_Type")).select_by_index("1")
            devicer.find_element(By.ID, "AnnexIV_Weight").send_keys(str(weight))
            devicer.find_element(By.ID, "AnnexIV_NoOfPerson").send_keys(int(Person_cap))
        else:
            Select(devicer.find_element(By.CSS_SELECTOR, "select#AnnexIV_Type")).select_by_index("2")
            devicer.find_element(By.ID, "AnnexIV_Weight").send_keys(str(weight))
            devicer.find_element(By.ID, "AnnexIV_NoOfPerson").send_keys(int(Person_cap))

        print("Selected Appliance Type.")
        devicer.save_screenshot("Make.png")

        print("Type and Make of Lift or Escalator Filled Successfully. Now,Filling the Manufacture Details.")
        devicer.find_element(By.ID, "AnnexIV_manufacturerName").click()

        devicer.find_element(By.ID, "AnnexIV_manufacturerName").send_keys(Make)
        time.sleep(2)
        devicer.find_element(By.ID, "AnnexIV_manufacturerName").send_keys(Keys.DOWN,Keys.ENTER)
        time.sleep(2)
        Name = devicer.find_element(By.ID, "AnnexIV_localAuthorizedManufacturerName").get_attribute("value")
        local_authorized_representative_number = devicer.find_element(By.ID, "AnnexIV_localAuthorizedManufactureContactDetail").get_attribute("value")
        print("Manufacturer Name:", Name)
        regi_no = devicer.find_element(By.ID, "AnnexIV_RegNoManufacturer").get_attribute("value")
        devicer.find_element(By.ID, "AnnexIV_agencyfacturerName").send_keys(Name)
        devicer.find_element(By.ID, "AnnexIV_RegNoAgency").send_keys(regi_no)
        devicer.find_element(By.ID, "AnnexIV_agency_HouseNo").send_keys(agency_house_no)
        devicer.find_element(By.ID, "AnnexIV_agency_Building").send_keys(agency_building)
        devicer.find_element(By.ID, "AnnexIV_agency_Landmark").send_keys(agency_landmark)
        devicer.find_element(By.ID, "AnnexIV_agency_Locality").send_keys(agency_locality)
        devicer.find_element(By.ID, "AnnexIV_agency_Pincode").send_keys(agency_pincode, Keys.TAB)
        time.sleep(2)
        devicer.find_element(By.ID, "AnnexIV_localAuthorizedagencyName").send_keys(local_authorized_representative)
        devicer.find_element(By.ID, "AnnexIV_localAuthorizedagencyContactDetail").send_keys(local_authorized_representative_number)
        devicer.find_element(By.ID, "AnnexIV_agency_Local_HouseNo").send_keys(agency_house_no)
        devicer.find_element(By.ID, "AnnexIV_agency_Local_Building").send_keys(agency_building)
        devicer.find_element(By.ID, "AnnexIV_agency_Local_Landmark").send_keys(agency_landmark)
        devicer.find_element(By.ID, "AnnexIV_agency_Local_Locality").send_keys(agency_locality)
        devicer.find_element(By.ID, "AnnexIV_agency_Local_Pincode").send_keys(agency_pincode, Keys.TAB)
        time.sleep(2)
        date1 = devicer.find_element(By.ID, "AnnexIV_Commencement_commissioning_Date")
        date2 = devicer.find_element(By.ID, "AnnexIV_Completion_commissioning_Date")
        devicer.execute_script("arguments[0].value = arguments[1];", date1, commence_date)
        devicer.execute_script("arguments[0].value = arguments[1];", date2, end_date)

        devicer.find_element(By.ID, "nxt4").click()
        print("Form 4 Filled Successfully.")
        devicer.save_screenshot("Step4.png")
        time.sleep(2)
        print("Now Uploading Files..")
        devicer.find_element(By.ID, "ApprovedbuildingPlan").send_keys(f"{OC_file}")
        time.sleep(3)
        element = devicer.find_element(By.XPATH, "/html/body/div[2]")
        devicer.execute_script("arguments[0].style.display= 'none';", element)
        time.sleep(2)
        devicer.find_element(By.ID, "DrawingdetailsoftheliftOrEscalator").send_keys(f'{Drawing}'),
        time.sleep(1)
        devicer.find_element(By.ID, "AffidavitOfManufacturer").send_keys(f'{Affidavit_Manufacturer}'),
        time.sleep(1)
        devicer.find_element(By.ID, "TechnicalDetails").send_keys(f'{Technical_detail}'),
        time.sleep(1)
        devicer.find_element(By.ID, "SafetyFeatures").send_keys(f'{Safety}'),
        time.sleep(1)
        devicer.find_element(By.ID, "SeparateDeclarations").send_keys(f'{Signed_by_three_Affidavit}'),
        time.sleep(1)
        devicer.find_element(By.ID, "ManuSignature").send_keys(f'{Manufacturer_Signature}'),
        time.sleep(1)
        devicer.find_element(By.ID, "AuthSig").send_keys(f'{Manufacturer_Signature}'),
        time.sleep(1)
        devicer.find_element(By.ID, "OSig").send_keys(f'{Owners_Signature}'),
        time.sleep(1)
        devicer.find_element(By.CSS_SELECTOR, "input#clari").click()
        time.sleep(3)
        devicer.find_element(By.ID, "nxt5").click()
        print("Form 6 also Filled Successfully.")
        devicer.save_screenshot('Step5.png')
        time.sleep(2)
        link = devicer.find_element(By.PARTIAL_LINK_TEXT, 'TEMP').text
        print(f'Your Registration No: {link}')
        file_path = "Registrations.xlsx"
        if os.path.exists(file_path):   
            wb = load_workbook(file_path)
            sheet = wb.active
        else:
            wb = Workbook()
            sheet = wb.active

        # Find the next empty row in column A
        row = sheet.max_row + 1
        sheet[f"A{row}"] = link

        # Adjust column width based on content
        text_length = len(str(link))
        sheet.column_dimensions["A"].width = max(sheet.column_dimensions["A"].width or 0, text_length + 5)

        # Save the file
        wb.save(file_path)
        print(f"Saved registration link in A{row}")

        devicer.save_screenshot(f"{link}.png")
        devicer.find_element(By.PARTIAL_LINK_TEXT, 'TEMP').click()
        devicer.find_element(By.ID, 'nxt1').click()
        time.sleep(1)
        devicer.find_element(By.ID, 'LiftEescalatorCommissionedPath').click()
        devicer.find_element(By.ID, 'PlaceOfCommissioningPath').click()
        devicer.find_element(By.ID, 'CommissioningAgencySamePath').click()
        devicer.find_element(By.ID, 'InCaseThereIsAnyChange').send_keys(Affidavit_Manufacturer)
        time.sleep(3)
        ani =  devicer.find_element(By.XPATH, '/html/body/div')
        devicer.execute_script('arguments[0].style.display= "none";', ani)
        time.sleep(2)
        devicer.find_element(By.ID, 'SelfDeclarationNotarizedAffidavit').send_keys(Affidavit_Manufacturer)
        time.sleep(1)
        devicer.find_element(By.ID, 'AffidavitCommissioningAgency').send_keys(Affidavit_Manufacturer)
        time.sleep(1)
        devicer.find_element(By.ID, 'AffidavitOfTheManufacturer').send_keys(Affidavit_Manufacturer)
        time.sleep(1)
        devicer.find_element(By.ID, 'SeparateDeclarationsOnnotarized').send_keys(Signed_by_three_Affidavit)
        time.sleep(1)
        input_field = devicer.find_element(By.ID, 'SuggestiveUsefulLife')
        input_field.click()
        input_field.send_keys('25')
        declar = devicer.find_element(By.CSS_SELECTOR, 'select#WhethertheOwnerOperatorOfTheLift')
        Select(declar).select_by_index(1)
        declar = devicer.find_element(By.ID, 'nxt2').click()
        print("All done!")
        devicer.save_screenshot("Final.png")
        time.sleep(2)
        forms+= 1
        print(f'Registration no: {forms} completed...')
        i +=1        
        if (i == how_many_form):
            break
login()
devicer.save_screenshot("filled_form.png")
devicer.quit()
